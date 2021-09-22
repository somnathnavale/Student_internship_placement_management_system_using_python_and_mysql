# -*- coding: utf-8 -*-
"""
Created on Wed Sep 22 13:31:10 2021

@author: somnath
"""
import openpyxl as o
import mysql.connector as c
con=c.connect(host="localhost",user="root",passwd="*******",database="placementportal")  #add your password and database  
cursor=con.cursor()


    
def write_into_sheet() :
    
    wb=o.load_workbook("form.xlsx")
    sheet1=wb["applied"]

    i=2                         #shows starting row from excel sheet
    lis=[]
    
    while(type(sheet1.cell(i,1).value)==int) :
        num=sheet1.cell(i,1).value
        lis.append(num)
        i=i+1
        
    selection(*lis)
    
    return


def insert_data():
    
    parameter=["gr_no", "fullname","gender-(Male/Female)","date_of_birth(YYYY-DD-MM)", "mobile_no(10 digit)", "alternate_mobile_no" ,"personal_email_id" ,"college_email_id" ,"tenth_percentage(% sign not allowed)" ,"twelth_or_diploma(12th/Diploma)" , "twelth_or_diploma_percentage" , "branch(B.tech E&TC/IT/Comp/Mechanical/Civil", "cgpa", "active_backlog" ,"dead_backlog" ,"internship_bit" ,"internship_and_ppo_bit" ,"placement_bit" ]
    int_parameter=["gr_no","mobile_no(10 digit)","alternate_mobile_no", "active_backlog" ,"dead_backlog" ,"internship_bit" ,"internship_and_ppo_bit" ,"placement_bit"]
    float_parameter=["tenth_percentage(% sign not allowed)","twelth_or_diploma_percentage","cgpa"]
   
    data=[]
    
    for i in range(0,len(parameter)):
        print("enter a ",parameter[i],end="  ");
        if parameter[i] in int_parameter :
            x=int(input())
            data.append(x)
        elif parameter[i] in float_parameter :
            x=float(input())
            data.append(x)
        else:
            x=input()
            data.append(x)
            
    
    query="INSERT INTO data values({},'{}','{}','{}',{},{},'{}','{}',{},'{}',{},'{}',{},{},{},{},{},{})".format(data[0],data[1],data[2],data[3],data[4],data[5],data[6],data[7],data[8],data[9],data[10],data[11],data[12],data[13],data[14],data[15],data[16],data[17])
    cursor.execute(query)
    con.commit()
    
    print("inserted successfully")
    ch=input("do you want to insert another record then press y else x\n")
    if ch!='y' :
        return
           

def update_data():
    while(True):
        gr_no=int(input("enter gr no\n"))
        choice=input("in which column you want to update internship_bit,internship_and_ppo_bit,placement_bit\n")
        val=int(input("enter 1 if selected else 0\n"))
        
        query="update data set {}={} where gr_no={}".format(choice,val,gr_no)
            
        cursor.execute(query)
        con.commit()
        print("updated successfully")
        
        ch=input("do you want to insert another record then press y else x\n")
        if ch!='y' :
            return
    
    
def selection(*lis) :
    
   # s=["tenth_percentage","twelth_or_percentage", "branch" ,"cgpa", "active_backlog" ,"dead_backlog" , "internship_bit","internship_and_ppo_bit","placement_bit"]
        
    tenth_percentage=int(input("enter 10th min percentage\n"))
    twelth_or_diploma_percentage=int(input("enter twelth or diploma min percentage\n"))
    
    branch=set();
    while(True) :
        choice =int(input("select branch 1:Comp 2:IT 3:E&TC 4:Mechanical 5:civil\n"))
        if choice==1 :
            branch.add("B.tech COMP")
        elif choice==2 :
            branch.add("B.tech IT")
        elif choice==3 :
            branch.add("B.tech E&TC")
        elif choice==4 :
            branch.add("B.tech Mechanical")
        elif choice==5 :
            branch.add("B.tech Civil")
        else:
            print("Wrong Choice")
        
        ch=input("enter Y to add one more Branch else enter x\n")
        if ch!='Y':
            break            
            
    cgpa=int(input("enter a minimum cgpa cutoff\n"))
    active_backlog=int(input("enter maximum allowed active backlogs\n"))
    dead_backlog=int(input("enter maximum allowed dead backlogs\n"))
    roll=input("select opportunity internship ,internship_and_ppo,placement\n")
    
    r=2
    wb=o.load_workbook("form.xlsx")
    sheet2=wb["shortlist"]
    
    for i in lis :
        
        query="select * from data where gr_no={}".format(i)
        cursor.execute(query)
        
        user_data=cursor.fetchall()
        bit=1
        
        if user_data[0][8]<tenth_percentage :
            bit=0
        if user_data[0][10]<twelth_or_diploma_percentage :
            bit=0
        if user_data[0][11] not in branch :
            bit=0
        if user_data[0][12] <cgpa :
            bit=0
        if user_data[0][13]>active_backlog:
            bit=0
        if user_data[0][14]>dead_backlog:
            bit=0
        if roll=="internship" and user_data[0][15]+user_data[0][16]==1:
            bit=0
        if roll=="internship_and_ppo" and user_data[0][15]+user_data[0][16]+user_data[0][17]>=1:
            bit=0
        if roll=="placement" and user_data[0][16]+user_data[0][17] >=1 :
            bit=0
        
        if bit==1 :
            
            for j in range(0,18) :          # 18 because we have 18 different student parametrs
                sheet2.cell(row=r,column=j+1,value=user_data[0][j])
            r=r+1
    
    wb.save("shortlisted.xlsx")
    return
 
if __name__=="__main__" :
    write_into_sheet()
    print("in func")    